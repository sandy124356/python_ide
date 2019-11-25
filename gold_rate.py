import lxml.html as LH
import requests
import ctypes
import xlwt
from datetime import datetime
import openpyxl


from xlwt import Workbook
#from xlutils.copy import copy
#import pandas as pd
def text(elt):
    return elt.text_content().replace(u'\xa0', u' ')

url = 'http://www.sify.com/finance/gold-rates-in-hyderabad/'
r = requests.get(url)
root = LH.fromstring(r.content)

print(type(root))

tr_elements = root.xpath('//*[@id="middle"]/div[2]/div[1]/div/div[2]/div/div/table/tbody/tr[1]/td[2]')

print(tr_elements)
#Create empty list
col=[]
i=0
#For each row, store each first element (header) and an empty list
for t in tr_elements:
    rate=t.text_content().replace(u'\xa0', u' ')
    print(type(rate))
    msg='1 Gram 22-Carat Gold Price (in HYD) today :'+rate
    print ('22-Carat Gold Price today: "%s"'%(msg))


ctypes.windll.user32.MessageBoxW(0, msg, "Alert", 0x1000)

# storing this value daily in a file

#wb = Workbook()

#sheet1 = wb.add_sheet('Gold_rates_history')

#sheet1.write(0, 0, name)

#wb.save('xlwt Gold.xls')



wb = openpyxl.load_workbook('Gold.xlsx')
sheet = wb.get_sheet_by_name('Gold_rates_history')

max_r=sheet.max_row
sheet['A' + str(max_r + 1)].value = datetime.today().strftime('%d-%m-%y')
sheet['B' + str(max_r + 1)].value = rate
print(sheet.max_row)
wb.save('Gold.xlsx')
print(sheet.max_row)
